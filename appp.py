import os
import re
import math
from datetime import datetime, date, time, timedelta
from io import BytesIO
from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
import pymysql
from pymysql.cursors import DictCursor
import pytz
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

# --- Configuration ---
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '@bdullah1266',   # CHANGE to your MySQL password
    'database': 'quastech_db',
    'charset': 'utf8mb4',
    'cursorclass': DictCursor,
    'autocommit': True
}

# College GPS coordinates (replace with your actual values)
COLLEGE_LAT = 19.148539
COLLEGE_LON = 73.036272
GPS_RADIUS_METERS = 200

IST = pytz.timezone('Asia/Kolkata')
ADMIN_CODE = 'admin1246'

def get_db_connection():
    try:
        return pymysql.connect(**DB_CONFIG)
    except pymysql.Error as e:
        print(f"❌ Database connection error: {e}")
        return None

# --- Password Functions (Plaintext) ---
def hash_password(password):
    return password  # plaintext

def check_password(password, stored):
    return password == stored

# --- Haversine Distance ---
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# --- Auto-Absent (10:01 PM IST) ---
def mark_absent_students():
    today = datetime.now(IST).date()
    conn = get_db_connection()
    if not conn:
        return
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT registration_number FROM registrations")
            all_students = [row['registration_number'] for row in cur.fetchall()]
            cur.execute("SELECT registration_number FROM attendance WHERE date = %s", (today,))
            present_today = [row['registration_number'] for row in cur.fetchall()]
            absent_students = set(all_students) - set(present_today)
            now_time = datetime.now(IST).time()
            for reg_num in absent_students:
                cur.execute(
                    "INSERT INTO attendance (registration_number, date, time_in, status, device_fingerprint) VALUES (%s, %s, %s, %s, %s)",
                    (reg_num, today, now_time, 'Absent', 'auto_absent')
                )
            conn.commit()
            print(f"✅ Marked {len(absent_students)} students as Absent for {today}")
    except Exception as e:
        print(f"❌ Auto-absent error: {e}")
    finally:
        conn.close()

# --- Helper: Convert timedelta to string ---
def timedelta_to_str(td):
    if td is None:
        return ''
    if isinstance(td, timedelta):
        total_seconds = int(td.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return str(td)

# --- Helper: Get month calendar data ---
def get_month_calendar(reg_number, year, month):
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT date, status FROM attendance 
                WHERE registration_number = %s AND YEAR(date) = %s AND MONTH(date) = %s
            """, (reg_number, year, month))
            records = cur.fetchall()
            attendance_map = {row['date'].day: row['status'] for row in records}
            return attendance_map
    finally:
        conn.close()

# --- Helper: Get dashboard stats ---
def get_dashboard_stats(reg_number):
    conn = get_db_connection()
    if not conn:
        return {'present': 0, 'absent': 0, 'working_days': 0}
    try:
        with conn.cursor() as cur:
            today = datetime.now(IST).date()
            first_day = today.replace(day=1)
            cur.execute("""
                SELECT COUNT(*) as total_working_days FROM attendance 
                WHERE registration_number = %s AND date BETWEEN %s AND %s
            """, (reg_number, first_day, today))
            total_days = cur.fetchone()
            
            cur.execute("""
                SELECT COUNT(CASE WHEN status IN ('Present', 'Late') THEN 1 END) as present,
                       COUNT(CASE WHEN status = 'Absent' THEN 1 END) as absent
                FROM attendance 
                WHERE registration_number = %s AND date BETWEEN %s AND %s
            """, (reg_number, first_day, today))
            stats = cur.fetchone()
            
            # Working days excluding Sundays
            working_days = 0
            for d in range((today - first_day).days + 1):
                current_date = first_day + timedelta(days=d)
                if current_date.weekday() != 6:
                    working_days += 1
            
            return {
                'present': stats.get('present', 0) if stats else 0,
                'absent': stats.get('absent', 0) if stats else 0,
                'working_days': working_days
            }
    finally:
        conn.close()

# --- Routes ---

@app.route('/')
def index():
    if session.get('student_logged_in'):
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        data = request.get_json()
        name = data.get('name', '').strip()
        reg_num = data.get('registration_number', '').strip()
        mobile = data.get('mobile', '').strip()
        year = data.get('year', '').strip()
        password = data.get('password', '').strip()

        # Admin shortcut
        if reg_num.lower() == ADMIN_CODE:
            session['admin_logged_in'] = True
            return jsonify({'success': True, 'redirect': '/admin', 'message': 'Admin access granted!'}), 200

        if not all([name, reg_num, mobile, year, password]):
            return jsonify({'success': False, 'message': 'All fields are required.'}), 400
        if len(mobile) != 10 or not mobile.isdigit():
            return jsonify({'success': False, 'message': 'Mobile must be exactly 10 digits.'}), 400
        if year not in ['FYBCA', 'SYBCA', 'TYBCA']:
            return jsonify({'success': False, 'message': 'Invalid year.'}), 400
        if len(password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM registrations WHERE registration_number = %s", (reg_num,))
                if cur.fetchone():
                    return jsonify({'success': False, 'message': 'Registration number already exists.'}), 400
                cur.execute("SELECT id FROM registrations WHERE mobile = %s", (mobile,))
                if cur.fetchone():
                    return jsonify({'success': False, 'message': 'Mobile number already registered.'}), 400

                # Store plaintext password
                cur.execute(
                    "INSERT INTO registrations (registration_number, name, mobile, year, password) VALUES (%s, %s, %s, %s, %s)",
                    (reg_num, name, mobile, year, password)
                )
                conn.commit()
                
                session['student_logged_in'] = True
                session['registration_number'] = reg_num
                session['student_name'] = name
                
                return jsonify({'success': True, 'redirect': '/dashboard', 'message': 'Signup successful!'})
        except pymysql.IntegrityError as e:
            return jsonify({'success': False, 'message': 'Database error: ' + str(e)}), 400
        finally:
            conn.close()
    
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        reg_num = data.get('registration_number', '').strip()
        password = data.get('password', '').strip()

        # 🚀 Admin shortcut – no password check
        if reg_num.lower() == ADMIN_CODE:
            session['admin_logged_in'] = True
            return jsonify({'success': True, 'redirect': '/admin', 'message': 'Admin access granted!'}), 200

        if not reg_num or not password:
            return jsonify({'success': False, 'message': 'Registration Number and Password are required.'}), 400

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT registration_number, name, password FROM registrations WHERE registration_number = %s", (reg_num,))
                student = cur.fetchone()
                if not student:
                    return jsonify({'success': False, 'message': 'Registration ID invalid.'}), 400
                # Plaintext comparison
                if student['password'] != password:
                    return jsonify({'success': False, 'message': 'Password invalid.'}), 400

                session['student_logged_in'] = True
                session['registration_number'] = student['registration_number']
                session['student_name'] = student['name']
                
                return jsonify({'success': True, 'redirect': '/dashboard', 'message': 'Login successful!'})
        finally:
            conn.close()
    
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    
    reg_number = session['registration_number']
    stats = get_dashboard_stats(reg_number)
    
    today = datetime.now(IST).date()
    current_month = today.month
    current_year = today.year
    
    return render_template('dashboard.html', 
                         student_name=session['student_name'],
                         reg_number=reg_number,
                         stats=stats,
                         current_month=current_month,
                         current_year=current_year)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/profile')
def profile():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT registration_number, name, mobile, year, course FROM registrations WHERE registration_number = %s", 
                       (session['registration_number'],))
            student = cur.fetchone()
            return render_template('profile.html', student=student)
    finally:
        conn.close()

@app.route('/calendar_data')
def calendar_data():
    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Not logged in.'}), 401
    
    year = request.args.get('year', type=int, default=datetime.now(IST).year)
    month = request.args.get('month', type=int, default=datetime.now(IST).month)
    
    reg_number = session['registration_number']
    attendance_map = get_month_calendar(reg_number, year, month)
    
    calendar_data = {}
    for day, status in attendance_map.items():
        calendar_data[str(day)] = status
    
    return jsonify({'success': True, 'data': calendar_data})

@app.route('/attendance_page')
def attendance_page():
    if not session.get('student_logged_in'):
        return redirect(url_for('login'))
    return render_template('attendance_form.html')

# --- Admin Routes ---

@app.route('/admin', methods=['GET', 'POST'])
def admin_panel():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == '1246':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_panel'))
        else:
            return render_template('admin.html', error="Wrong password")
    if not session.get('admin_logged_in'):
        return render_template('admin.html', error=None)
    
    conn = get_db_connection()
    registrations = []
    total_registrations = 0
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM registrations ORDER BY registered_at DESC")
                registrations = cur.fetchall()
                total_registrations = len(registrations)
        finally:
            conn.close()
    
    return render_template('admin.html', registrations=registrations, total_registrations=total_registrations)

@app.route('/admin/search')
def admin_search():
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'Not authorized.'}), 401
    
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({'success': True, 'data': []})
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT * FROM registrations 
                WHERE name LIKE %s OR registration_number LIKE %s
                ORDER BY registered_at DESC
            """, (f'%{query}%', f'%{query}%'))
            results = cur.fetchall()
            return jsonify({'success': True, 'data': results})
    finally:
        conn.close()

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('index'))

# --- Attendance API (modified) ---

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    data = request.get_json()
    registration_number = data.get('registration_number', '').strip()
    year = data.get('year', '').strip()
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    device_fingerprint = data.get('device_fingerprint', '').strip()

    # Admin shortcut
    if registration_number.lower() == ADMIN_CODE:
        session['admin_logged_in'] = True
        return jsonify({'success': True, 'redirect': '/admin', 'message': 'Admin access granted!'}), 200

    if not session.get('student_logged_in'):
        return jsonify({'success': False, 'message': 'Please login first.'}), 401

    if registration_number != session['registration_number']:
        return jsonify({'success': False, 'message': 'You can only mark your own attendance.'}), 400

    if not all([registration_number, year]):
        return jsonify({'success': False, 'message': 'Registration Number and Year are required.'}), 400
    if year not in ['FYBCA', 'SYBCA', 'TYBCA']:
        return jsonify({'success': False, 'message': 'Invalid year.'}), 400
    if latitude is None or longitude is None:
        return jsonify({'success': False, 'message': 'Location access is required.'}), 400
    if not device_fingerprint:
        return jsonify({'success': False, 'message': 'Device fingerprint not available.'}), 400

    distance = haversine(COLLEGE_LAT, COLLEGE_LON, float(latitude), float(longitude))
    if distance > GPS_RADIUS_METERS:
        return jsonify({'success': False, 'message': f'You are not within college campus. Distance: {int(distance)} meters.'}), 400

    now = datetime.now(IST)
    current_time = now.time()
    current_date = now.date()

    if current_time < time(12, 0):
        return jsonify({'success': False, 'message': 'Attendance window starts at 12:00 PM IST.'}), 400
    if current_time >= time(22, 0):
        return jsonify({'success': False, 'message': 'Attendance window closed. Timeout.'}), 400

    status = 'Present' if current_time < time(13, 0) else 'Late'

    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Database connection failed.'}), 500
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT year FROM registrations WHERE registration_number = %s", (registration_number,))
            student = cur.fetchone()
            if not student:
                return jsonify({'success': False, 'message': 'Invalid Registration Number.'}), 400
            if student['year'] != year:
                return jsonify({'success': False, 'message': f"Student is in {student['year']}, not {year}."}), 400

            cur.execute("SELECT id FROM attendance WHERE registration_number = %s AND date = %s", (registration_number, current_date))
            if cur.fetchone():
                return jsonify({'success': False, 'message': 'Attendance already marked for today.'}), 400

            cur.execute("SELECT registration_number FROM attendance WHERE device_fingerprint = %s AND date = %s", (device_fingerprint, current_date))
            existing = cur.fetchone()
            if existing and existing['registration_number'] != registration_number:
                return jsonify({'success': False, 'message': 'This device has already been used for attendance today.'}), 400

            cur.execute(
                "INSERT INTO attendance (registration_number, date, time_in, status, device_fingerprint) VALUES (%s, %s, %s, %s, %s)",
                (registration_number, current_date, current_time, status, device_fingerprint)
            )
            conn.commit()
            return jsonify({'success': True, 'message': f'Attendance marked as {status}.', 'status': status})
    finally:
        conn.close()

# --- Download Report (3 Sheets) ---
@app.route('/download_full_report', methods=['GET'])
def download_full_report():
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ---- Sheet 1: Registrations ----
    ws1 = wb.active
    ws1.title = "Registrations"
    conn = get_db_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT registration_number, name, mobile, year, course, password, registered_at FROM registrations ORDER BY registration_number")
                regs = cur.fetchall()
                if regs:
                    ws1.append(['Registration Number', 'Name', 'Mobile', 'Year', 'Course', 'Password', 'Registered At'])
                    for reg in regs:
                        ws1.append([
                            reg['registration_number'],
                            reg['name'],
                            reg['mobile'],
                            reg['year'],
                            reg.get('course', 'BCA'),
                            reg['password'],
                            reg['registered_at'].strftime('%Y-%m-%d %H:%M:%S') if reg['registered_at'] else ''
                        ])
                else:
                    ws1.append(['No registrations found.'])
        finally:
            conn.close()

    # ---- Sheet 2: Attendance ----
    ws2 = wb.create_sheet("Attendance")
    conn = get_db_connection()
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT a.registration_number, r.name, a.date, a.time_in, a.status
                    FROM attendance a
                    JOIN registrations r ON a.registration_number = r.registration_number
                    ORDER BY a.date DESC, a.registration_number
                """)
                atts = cur.fetchall()
                if atts:
                    ws2.append(['Registration Number', 'Name', 'Date', 'Time In', 'Status'])
                    for att in atts:
                        time_str = timedelta_to_str(att.get('time_in'))
                        ws2.append([
                            att['registration_number'],
                            att['name'],
                            att['date'].strftime('%Y-%m-%d') if att['date'] else '',
                            time_str,
                            att['status']
                        ])
                else:
                    ws2.append(['No attendance records found.'])
        finally:
            conn.close()

    # ---- Sheet 3: Student Slicer ----
    ws3 = wb.create_sheet("Student Slicer")
    now = datetime.now(IST)
    current_month = now.month
    current_year = now.year
    conn = get_db_connection()
    data_rows = []
    if conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT a.registration_number, r.name, a.date, a.status,
                           DAY(a.date) as day_of_month
                    FROM attendance a
                    JOIN registrations r ON a.registration_number = r.registration_number
                    WHERE MONTH(a.date) = %s AND YEAR(a.date) = %s
                    ORDER BY a.registration_number, a.date
                """, (current_month, current_year))
                records = cur.fetchall()
                student_weekly = {}
                for rec in records:
                    reg_num = rec['registration_number']
                    name = rec['name']
                    day = rec['day_of_month']
                    if day <= 7:
                        week = 1
                    elif day <= 14:
                        week = 2
                    elif day <= 21:
                        week = 3
                    else:
                        week = 4
                    status = rec['status']
                    key = (reg_num, week)
                    if key not in student_weekly:
                        student_weekly[key] = {'Present':0, 'Late':0, 'Absent':0, 'Timeout':0, 'total_days':0, 'name':name}
                    student_weekly[key][status] += 1
                    student_weekly[key]['total_days'] += 1
                for (reg_num, week), counts in student_weekly.items():
                    total = counts['total_days']
                    perc = round((counts['Present'] + counts['Late']) / total * 100, 2) if total > 0 else 0
                    data_rows.append([
                        reg_num,
                        counts['name'],
                        week,
                        counts['Present'],
                        counts['Late'],
                        counts['Absent'],
                        counts['Timeout'],
                        total,
                        perc
                    ])
        finally:
            conn.close()

    wb.create_sheet("WeeklyData")
    ws_data = wb["WeeklyData"]
    ws_data.sheet_state = 'hidden'
    ws_data.append(['Reg Number', 'Name', 'Week', 'Present', 'Late', 'Absent', 'Timeout', 'Total Days', 'Attendance %'])
    for row in data_rows:
        ws_data.append(row)

    ws3.title = "Student Slicer"
    student_names = sorted(set(row[1] for row in data_rows))
    if not student_names:
        ws3.append(['No data for this month.'])
    else:
        dv = DataValidation(type="list", formula1='"{}"'.format(','.join(student_names)))
        ws3.add_data_validation(dv)
        dv.add('B1')
        ws3['A1'] = "Select Student:"
        ws3['B1'] = student_names[0]

        ws3['A3'] = "Week"
        ws3['B3'] = "Attendance %"
        for i in range(1, 5):
            ws3.cell(row=3+i, column=1, value=f"Week {i}")
            ws3.cell(row=3+i, column=2).value = f'=SUMIFS(WeeklyData!I:I, WeeklyData!B:B, $B$1, WeeklyData!C:C, {i})'

        chart_data = Reference(ws3, min_col=2, min_row=4, max_row=7)
        categories = Reference(ws3, min_col=1, min_row=4, max_row=7)
        chart = BarChart()
        chart.title = "Weekly Attendance Percentage"
        chart.x_axis.title = "Week"
        chart.y_axis.title = "Percentage"
        chart.add_data(chart_data, titles_from_data=False)
        chart.set_categories(categories)
        ws3.add_chart(chart, "D3")

        for row in ws3.iter_rows(min_row=3, max_row=7, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    filename = "attendance_report.xlsx"
    return send_file(
        file_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- Scheduler ---
scheduler = BackgroundScheduler()
scheduler.add_job(
    mark_absent_students,
    'cron',
    hour=22,
    minute=1,
    timezone=IST
)
scheduler.start()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)